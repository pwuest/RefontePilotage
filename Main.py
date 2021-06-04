import pandas as pd
import numpy as np
import openpyxl as pyxl

from ExcelLoad import Load_Table_From_Excel, Clean_Table, First_Row_Header, DataFrame_toStr
from ProduitCartésien import select_avant_realloc, cartesian_product_multi

folder = "C:/Users/admin/Documents/Pilotage LBP"
path_referentiel = folder + "/0. Référentiel des données - 020.xlsx"
path_excel = folder + "/Cartographie - Charges - 005.xlsx"
# path_excel = folder + "/Cartographie - Charges - 002.xlsx"
separator_csv = ";"
suffixFichier = "_Charges"

referentiel_address_dict = {'Indicateur': ["Nouveau_REF_Conso", "E4:I"],
                            'Produit': ["Produits", "D8:H"],
                            'Entité juridique': ["EJ", "D8:O"],
                            'Unité de gestion / Direction': ["UG", "D8:I"],
                            'Distribution': ["Distribution", "D8:G"],
                            'Segment client / marché': ["Clients", "D8:H"]}


# Load référentiel ----------------------------------------------------------------------------------------------------
wb = pyxl.load_workbook(path_referentiel, read_only=True, keep_vba=False, data_only=True, keep_links=False)
Referentiel = Load_Table_From_Excel(referentiel_address_dict, wb, bFirstRowasHeader=True)
wb.close()

Referentiel['Contrepartie'] = Referentiel['Unité de gestion / Direction']


# Load Tableau de bord ------------------------------------------------------------------------------------------------
wb = pyxl.load_workbook(path_excel, read_only=True, keep_vba=False, data_only=True, keep_links=False)

table_address_dict = dict()
for name in wb.sheetnames:
    if ((not name[0].isdigit()) and name[1] == "." and name[2].isdigit()):
        if name[0] == "B.1 Charges métiers": #  "B.7. DEDT - Charges": #"C.1.1.a. PPH dashboard":
            table_address_dict[name] = [name, ""]

table_address_dict["Ratios"] = ["Ratios", "A1:P"]
table_address_dict["Exclusions"] = ["Exclusions", "A1:"]

Table = Load_Table_From_Excel(table_address_dict, wb)

Table["Exclusions"] = First_Row_Header(Table["Exclusions"])
Table["Ratios"] = DataFrame_toStr((First_Row_Header(Table["Ratios"])))
wb.close()

# -------------------------------------------------------------------------------------------------------------------- #
# 1ere étape - Rassembler le tout en une seule
# -------------------------------------------------------------------------------------------------------------------- #
Out = pd.DataFrame()
for table_name in Table.keys():
    if table_name not in ["Ratios", "Exclusions"]:
        print(table_name)
        Out = pd.concat([Out, Clean_Table(Table[table_name], table_name)])

Out.columns = ["Entité juridique", "Unité de gestion / Direction", "Contrepartie", "Segment client / marché", "Distribution", "Produit", "Indicateur", "Attribut", "Phase", "Adresse"]
Table["Ratios"].columns = ["_" + col for col in Out.columns.tolist()[0:8]] + Out.columns.tolist()[0:8]

Out = DataFrame_toStr(Out).astype(str)
# Out.to_csv(folder + "/Carto_out" + suffixFichier + ".csv", encoding='utf-8-sig', index=False, sep=separator_csv)
Out.to_excel(folder + "/Carto_out" + suffixFichier + ".xlsx", encoding='utf-8-sig', index=False)
Out_original = Out.copy()

# -------------------------------------------------------------------------------------------------------------------- #
# 2eme étape - Merger avec le référentiel pour décomposer en indicateurs élémentaires
# -------------------------------------------------------------------------------------------------------------------- #

# Ne faire qu'une colonne sur l'unité de gestion : si avant réallocation non vide alors prendre celle-là
Out["Contrepartie"] = Out.apply(lambda row: select_avant_realloc(row), axis=1)
Out.drop(["Unité de gestion / Direction"], axis=1, inplace=True)
Out.rename(columns={"Contrepartie": "Unité de gestion / Direction"})

Ref_columns = [col for sublist in [Referentiel[dim].columns.tolist() for dim in Out.columns[0:6]] for col in sublist]

# On éclate en indicateurs élémentaires
Out_eclate = dict()
Out_iter = pd.concat([Table["Ratios"][Out.columns[0:7]], Out[Out.columns[0:7]]])
Out_iter = Out_iter.drop_duplicates().reset_index(drop=True)
for index, row in Out_iter.iterrows():
    print(str(index) + "/" + str(Out_iter.shape[0]-1))
    row_ = pd.DataFrame(row[:-1]).transpose()
    Out_ = row_.copy()
    if not row["Attribut"] in ["Ratio", "Total"]:
        for dimension in row_.columns:
            item = row[dimension]
            if item == 'Tous':
                match_ = Referentiel[dimension]
            else:
                match_ = Referentiel[dimension][Referentiel[dimension].isin([item]).any(axis=1)].copy()

            if match_.shape[0] == 0 or Out_.shape[0] == 0:
                # Cartesian product
                Out_['key'] = 0
                match_['key'] = 0
                Out_ = Out_.merge(match_, on='key', how='outer').drop(columns=['key'])
            else:
                Out_ = cartesian_product_multi(*[Out_, match_])

        Out_.columns = [col for col in row_.columns.tolist() if col != 'key'] + Ref_columns
        Out_eclate[index] = Out_

# On rassemble les bases éclatées
Base = pd.concat([v for k, v in Out_eclate.items()], axis=0)

# Vision plus ramassée
col_not0 = [col for col in Base.columns.tolist() if (col in Out.columns.tolist() or (col[-2:] not in ['N0', 'N1']))]
Base_not0 = Base[col_not0].drop_duplicates()

# -------------------------------------------------------------------------------------------------------------------- #
# 3eme étape - On rassemble avec la base initiale sans certains niveaux
# -------------------------------------------------------------------------------------------------------------------- #
Base_ratio = pd.merge(Out_original[Out_original["Attribut"].isin(["Ratio"])], Table["Ratios"],
                      left_on=Base_not0.columns[0:6].tolist(),
                      right_on= ["_" + col for col in Base_not0.columns[0:6].tolist()], how='left', suffixes=('', '_y'))

temp_ratio = pd.merge(Base_ratio, Base_not0,
                      left_on=[col + '_y' for col in Base_not0.columns[0:6].tolist()],
                      right_on=Base_not0.columns[0:6].tolist(), how='left', suffixes=('', '_z'))

Base_complete = pd.merge(Base_not0, Out_original[~Out_original["Attribut"].isin(["Ratio", "Total"])],
                         on=Base_not0.columns[0:6].tolist(), how='left').dropna(subset=['Adresse'])

# Base_complete = pd.concat([Base_complete, Out_original[Out_original["Type d'indicateur"].isin(["Ratio", "Total"])]])
Base_complete = pd.concat([Base_complete, temp_ratio], axis=0, join='inner', ignore_index=True)
Base_complete = Base_complete.drop(["Phase"], axis=1).reset_index(drop=True)

Base_complete = Base_complete.astype(object).replace(np.nan, 'None')
Base_complete = Base_complete.astype(str)

# Version où l'on rassemble les adresses ensemble
Base_dimensions = Base_complete.drop(columns=Out.columns.tolist()[:-2])
Base_dimensions = Base_complete.groupby([col for col in Base_complete.columns if not col in Out.columns])['Adresse'].apply('/'.join).reset_index()

# On exclut certaines lignes
for index, row in Table["Exclusions"][[col for col in Table["Exclusions"].columns
                                       if col in Base_dimensions.columns]].iterrows():
    row_ = pd.DataFrame(row).transpose().dropna(axis=1, how='all')
    if row_.shape[0] > 0 and row_.shape[1] > 0:
        outer_join = Base_dimensions.merge(row_, how='outer', indicator=True)
        Base_dimensions = outer_join[outer_join._merge == 'left_only'].drop('_merge', axis=1)

# Eliminate duplicates
col_unique = ['Entite N2', 'UG N2', 'Segment N2', 'Distribution N2', 'Produit N2', 'Indicateur N2', 'Unité de gestion / Direction', 'Adresse']
Base_dimensions = Base_dimensions[col_unique].drop_duplicates()

# Write output
Base_dimensions.replace(["None", "nan"], np.nan, inplace=True)
# Base_dimensions.to_csv(folder + "/Base_dimensions" + suffixFichier + ".csv", encoding='utf-8-sig', index=False, sep=separator_csv)
Base_dimensions.to_excel(folder + "/Base_dimensions" + suffixFichier + ".xlsx", encoding='utf-8-sig', index=False)
